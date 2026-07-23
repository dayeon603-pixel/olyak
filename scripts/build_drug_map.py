"""약품명 → 주성분코드 매핑 파이프라인 (올약).

심평원(건강보험심사평가원) 「약제급여목록 및 급여 상한금액표」는 매월 1일 갱신되며
제품명과 주성분코드를 함께 담고 있다. 이 스크립트는 그 목록표를 읽어
제품명 ↔ 주성분코드 매핑 테이블(JSON)을 생성하고, OCR·검색으로 들어온
자유 텍스트 약품명을 주성분코드로 잇는 정규화·퍼지 매칭을 제공한다.

설계 원칙(남경제 교수 자문 2026-07-23 반영):
  - 매핑을 자체 사전으로 구축하지 않고, 공식 약제급여목록표를 "매월 동기화"한다.
  - 새 약이 나오면 목록표가 갱신되므로 매달 재생성하면 최신 상태가 유지된다.

데이터 출처:
  건강보험심사평가원 약제급여목록표 (매월 1일 갱신).
  다운로드 경로/공공데이터 API는 README(scripts/README.md) 참고.

의존성: 표준 라이브러리만 사용(CSV). .xlsx 입력은 openpyxl이 있으면 지원.

사용법:
  python build_drug_map.py --input 약제급여목록.csv --out ../data/drug_map.json
  python build_drug_map.py --input 약제급여목록.csv --lookup "노바스크정5mg"
"""
from __future__ import annotations

import argparse
import csv
import json
import logging
import re
import sys
from dataclasses import dataclass, asdict
from difflib import get_close_matches
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
logger = logging.getLogger("build_drug_map")

# 약제급여목록표에서 흔히 쓰이는 컬럼명 후보 (실제 파일에 맞춰 --name-col/--code-col로 덮어쓰기)
PRODUCT_COL_CANDIDATES = ("제품명", "품명", "약품명", "상품명")
CODE_COL_CANDIDATES = ("주성분코드", "성분코드", "일반명코드")
INGREDIENT_COL_CANDIDATES = ("주성분", "성분명", "일반명")

# 제품명 정규화 시 제거하는 제형/용량 꼬리표 (매칭 정확도용)
_FORM_SUFFIX = re.compile(
    r"(정|캡슐|캅셀|주|시럽|산|현탁액|점안액|연고|크림|패치|서방정|장용정|구강붕해정)?\s*$"
)
_DOSE = re.compile(r"\d+(\.\d+)?\s*(mg|밀리그람|g|㎎|mcg|㎍|iu|%)", re.IGNORECASE)
_PAREN = re.compile(r"[\(（].*?[\)）]")
_NONWORD = re.compile(r"[\s\.\-_/]+")


@dataclass(slots=True)
class DrugEntry:
    """매핑 한 건: 제품명 → 주성분코드(+성분명)."""

    product: str
    ingredient_code: str
    ingredient: str | None = None


def normalize(name: str) -> str:
    """제품명을 매칭용 키로 정규화한다.

    Args:
        name: 원본 제품명(또는 OCR/검색 자유 텍스트).
    Returns:
        용량·괄호·제형·공백을 제거한 소문자 키.
    """
    s = _PAREN.sub("", name)
    s = _DOSE.sub("", s)
    s = _NONWORD.sub("", s)
    s = _FORM_SUFFIX.sub("", s)
    return s.strip().lower()


def _pick_column(header: list[str], candidates: tuple[str, ...], override: str | None) -> str:
    """헤더에서 후보 컬럼명 중 하나를 고른다.

    Raises:
        KeyError: 후보 컬럼이 없고 override도 지정되지 않은 경우.
    """
    if override:
        if override not in header:
            raise KeyError(f"지정한 컬럼 '{override}'가 헤더에 없음: {header}")
        return override
    for c in candidates:
        if c in header:
            return c
    raise KeyError(f"컬럼 후보 {candidates} 중 헤더에 없음: {header}. --*-col로 지정하세요.")


def _read_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    """CSV 또는 XLSX에서 (헤더, 행 목록)을 읽는다."""
    if path.suffix.lower() in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError as e:  # pragma: no cover
            raise SystemExit("xlsx 입력에는 openpyxl이 필요합니다: pip install openpyxl") from e
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = [str(c).strip() if c is not None else "" for c in next(rows)]
        data = [
            {header[i]: ("" if v is None else str(v).strip()) for i, v in enumerate(r)}
            for r in rows
        ]
        return header, data
    # CSV (인코딩은 cp949/utf-8-sig 순차 시도 — 심평원 파일은 cp949가 흔함)
    for enc in ("utf-8-sig", "cp949", "utf-8"):
        try:
            text = path.read_text(encoding=enc)
            break
        except UnicodeDecodeError:
            continue
    else:  # pragma: no cover
        raise SystemExit(f"인코딩을 감지하지 못함: {path}")
    reader = csv.DictReader(text.splitlines())
    header = list(reader.fieldnames or [])
    return header, list(reader)


def build_map(
    path: Path,
    name_col: str | None = None,
    code_col: str | None = None,
    ing_col: str | None = None,
) -> dict[str, DrugEntry]:
    """약제급여목록표에서 정규화 제품명 → DrugEntry 매핑을 만든다."""
    header, rows = _read_rows(path)
    pcol = _pick_column(header, PRODUCT_COL_CANDIDATES, name_col)
    ccol = _pick_column(header, CODE_COL_CANDIDATES, code_col)
    try:
        icol: str | None = _pick_column(header, INGREDIENT_COL_CANDIDATES, ing_col)
    except KeyError:
        icol = None  # 성분명은 선택 사항

    mapping: dict[str, DrugEntry] = {}
    skipped = 0
    for row in rows:
        product = (row.get(pcol) or "").strip()
        code = (row.get(ccol) or "").strip()
        if not product or not code:
            skipped += 1
            continue
        key = normalize(product)
        if not key:
            skipped += 1
            continue
        mapping[key] = DrugEntry(
            product=product,
            ingredient_code=code,
            ingredient=(row.get(icol).strip() if icol and row.get(icol) else None),
        )
    logger.info("매핑 %d건 생성 (건너뜀 %d, 컬럼: %s/%s)", len(mapping), skipped, pcol, ccol)
    return mapping


def lookup(mapping: dict[str, DrugEntry], query: str, cutoff: float = 0.82) -> DrugEntry | None:
    """자유 텍스트 약품명을 매핑에서 찾는다(정규화 후 완전일치 → 퍼지).

    Args:
        mapping: build_map 결과.
        query: OCR/검색 자유 텍스트.
        cutoff: 퍼지 매칭 최소 유사도(0~1).
    Returns:
        가장 근접한 DrugEntry, 없으면 None.
    """
    key = normalize(query)
    if key in mapping:
        return mapping[key]
    near = get_close_matches(key, mapping.keys(), n=1, cutoff=cutoff)
    return mapping[near[0]] if near else None


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="약제급여목록표 → 주성분코드 매핑 생성")
    ap.add_argument("--input", type=Path, required=True, help="약제급여목록표 (.csv/.xlsx)")
    ap.add_argument("--out", type=Path, help="매핑 JSON 출력 경로")
    ap.add_argument("--name-col", help="제품명 컬럼명(자동감지 실패 시)")
    ap.add_argument("--code-col", help="주성분코드 컬럼명(자동감지 실패 시)")
    ap.add_argument("--ing-col", help="성분명 컬럼명(선택)")
    ap.add_argument("--lookup", help="테스트: 이 약품명을 매핑에서 조회")
    args = ap.parse_args(argv)

    if not args.input.exists():
        logger.error("입력 파일 없음: %s", args.input)
        return 1

    mapping = build_map(args.input, args.name_col, args.code_col, args.ing_col)

    if args.lookup:
        hit = lookup(mapping, args.lookup)
        print(json.dumps(asdict(hit) if hit else None, ensure_ascii=False, indent=2))

    if args.out:
        args.out.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "source": "심평원 약제급여목록표 (매월 1일 갱신)",
            "generated_from": str(args.input.name),
            "count": len(mapping),
            "map": {k: asdict(v) for k, v in mapping.items()},
        }
        args.out.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        logger.info("저장: %s (%d건)", args.out, len(mapping))

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
